# Databricks notebook source
# MAGIC %run ../utils

# COMMAND ----------

# MAGIC %pip install xlsxwriter

# COMMAND ----------

import tempfile
import shutil

from pyspark.sql.functions import col, sum as _sum, lit
from pyspark.sql import functions as F

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

import logging
logger = logging.getLogger()  # Get the root logger
logging.getLogger("py4j").setLevel(logging.WARNING)
logger.setLevel(logging.INFO)

PUBLISH_WITH_BOOST = True
APPLY_BLUE_FONT_IF_MISSING = False

OUTPUT_DIR = f"{TOP_DIR}/Workspace/output_excel"
INPUT_AUXI_DIR = f"{TOP_DIR}/Documents/input/Auxiliary"

CCI_FILE_PATH = f"{TOP_DIR}/Workspace/cci_csv/ALB/Executed.csv"

OUTPUT_FILE_PATH = f"{OUTPUT_DIR}/Albania BOOST.xlsx"
SOURCE_FILE_PATH = f"{INPUT_DIR}/Albania BOOST.xlsx"
TARGET_TABLE = 'prd_mega.boost_intermediate.alb_publish'
TARGET_TABLE_REVENUE = 'prd_mega.boost_intermediate.alb_boost_rev_gold'
OUTPUT_MISSING_DESC_FILE_PATH = f"{OUTPUT_DIR}/Albania_missing_code_descriptions.xlsx"



# COMMAND ----------

# Load and filter
raw_data = (
    spark.table(f"{TARGET_TABLE}")
    .withColumn("year", col("year").cast("string"))
    .cache()
)

tag_code_mapping =  pd.read_csv(TAG_MAPPING_URL)
years = [str(year) for year in sorted(raw_data.select("year").distinct().rdd.flatMap(lambda x: x).collect())]

def create_pivot(df, parent, child, agg_col ):
    filtered_mapping = tag_code_mapping[~tag_code_mapping['subnational']]

    # Step 1: Get detailed level e.g. (econ + econ_sub + year)
    detailed = (
        df.groupBy(parent, child, "year")
        .agg(F.sum(agg_col).alias(agg_col))
        .withColumnRenamed(parent, "parent")
        .withColumnRenamed(child, "child")
    )

    # Step 2: Get subtotals at parent level e.g (econ + year), econ_sub = 'Subtotal'
    subtotals = (
        df.groupBy(parent, "year")
        .agg(F.sum(agg_col).alias(agg_col))
        .withColumn("child", F.lit("Subtotal"))  # Ensure same schema
        .withColumnRenamed(parent, "parent")
    )

    # Step 3: Union both
    combined = detailed.unionByName(subtotals).filter(F.col("year").isNotNull())

    # Step 4: Pivot to wide format
    pivoted = (
        combined.groupBy("parent", "child")
        .pivot("year")
        .agg(F.sum(agg_col))
        .fillna(0)  # Replace NaNs with 0
    )

    # Step 5: Join with filtered_mapping
    filtered_mapping_spark = spark.createDataFrame(filtered_mapping)
    result = (
        pivoted.join(filtered_mapping_spark, on=["parent", "child"])
        .drop("category", "parent", "child", "parent_type", "child_type", "subnational")
    )

    total_matches = result.select("code").distinct().rdd.flatMap(lambda x: x).collect()
    logging.info(f"Matched {agg_col} entries for {parent}, {child} : {len(total_matches)} ")

    return result


# COMMAND ----------

def create_pivot_total(df, agg_col):
    # Step 1: Calculate total expenditure by year
    total = (
        df.groupBy("year")
        .agg(F.sum(agg_col).alias(agg_col))
        .withColumn("code", F.lit("EXP_ECON_TOT_EXP_EXE"))
    )

    # Step 2: Calculate foreign expenditure by year
    foreign_total = (
        df.filter(F.col("is_foreign") == True)
        .groupBy("year")
        .agg(F.sum(agg_col).alias(agg_col))
        .withColumn("code", F.lit("EXP_ECON_TOT_EXP_FOR_EXE"))
    )

    # Step 3: Combine total and foreign expenditure
    combined = total.unionByName(foreign_total).filter(F.col("year").isNotNull())

    # Step 4: Pivot to wide format
    pivoted = (
        combined.groupBy("code")
        .pivot("year")
        .agg(F.first(agg_col))
        .fillna(0)  
    )
    return pivoted

# COMMAND ----------

pairs = [
    ("boost_econ", "boost_econ_sub"),
    ("boost_func", "boost_econ_sub"),
    ("boost_func", "boost_econ"),
    ("boost_func", "boost_func_sub"),
    ("boost_func_sub", "boost_econ"),
    ("boost_func_sub", "boost_econ_sub"),
]

filtered_raw_data = raw_data.filter(col('transfer') == 'Excluding transfers')

def generate_combined_pivots(pairs, agg_col):
    # Initialize an empty DataFrame for combining results
    combined = None

    for parent, child in pairs:
        # Create pivot for central and regional levels
        pivoted = create_pivot(filtered_raw_data, parent, child, agg_col)

        # Combine the results
        if combined is None:
            combined = pivoted
        else:
            combined = combined.unionByName(pivoted)

    # Add totals to the combined DataFrame
    totals = create_pivot_total(filtered_raw_data, agg_col)
    combined = combined.unionByName(totals)
    return combined


executed = generate_combined_pivots(pairs, "executed")
approved = generate_combined_pivots(pairs, "approved")

# COMMAND ----------

# Helper functions to update the EXCEL sheets
#todo: move to utils for reusability when we have more than one country

def get_latest_cci_year(cci_df):
    year_columns = [col for col in cci_df.columns if col.split(".")[0].isdigit()]
    year_columns.sort(reverse=True)
    for year in year_columns:
        if not pd.isnull(cci_df[cci_df.Code == "EXP_ECON_TOT_EXP_EXE"][year].values[0]):
            return int(float(year))

# download the executed sheet from cci_csv to obtain the column list. 
template = pd.read_csv(CCI_FILE_PATH, dtype="str")
EXECUTED_TEMP_COL_LIST = [
   col.split(".")[0] for col in template.columns
]
BOOST_LATEST_YEAR  = int(max(years))
CCI_LATEST_YEAR = get_latest_cci_year(template)

def spark_to_pandas_with_reorder(ws, raw_data, include_boost_col=True):
    # make sure that the data expendture column specific order so that the formula will work
    raw_data = raw_data.toPandas()
    column_names = [cell.value for cell in ws[1] if cell.value is not None]  # Row 1 is typically the header

    remaining = [col for col in raw_data.columns if col not in column_names] if include_boost_col else []
    new_order = column_names + remaining
    for col in column_names:
        if col not in raw_data.columns:
            raw_data[col] = None
    raw_data = raw_data[new_order]
    return raw_data

def copy_font(cell, blue_text_format=False):
    if not cell.font:
        return {}
    # Copy font formatting
    bold = True if cell.font.bold else False
    italic = True if cell.font.italic else False
    num_format = cell.number_format
    font_size = cell.font.size if cell.font.size else 10  # Default font size if none specified
    font_name = cell.font.name if cell.font.name else 'Arial'  # Default to Arial if no font specified

    # source worksheet does not have other styling information. use heuristic to set the color/alignment. 
    if cell.row == 1:
        font_color = "#FFFFFF"
        bg_color = "#4d93d9"
    else:
        font_color = 'blue' if blue_text_format else '#000000'
        bg_color = "#FFFFFF"
    
    if cell.column < 3:
        align = 'left'
    else:
        align = 'right'

    return {
        'bold': bold,
        'italic': italic,
        'font_name': font_name,
        'font_size': font_size,
        'align': align,
        'valign': 'vcenter',
        'num_format': num_format,
        'font_color':font_color,
        "bg_color": bg_color,
    }

def set_width(target_ws,max_col_index):
    for i in range(max_col_index):
        if i == 0:
            width = 30
        elif i == 1:
            width = 50
        else:
            width = 12
        target_ws.set_column(i, i, width)  # xlsxwriter uses 0-based index

def get_col_name(col_inedex):
    # Some year columns in the original file use formulas (e.g., =U1+1),
    # which makes evaluating the actual column names dynamically too costly.
    # As a workaround, we get the list of col names from the cci_csv file.
    col_name = EXECUTED_TEMP_COL_LIST[col_inedex]
    return col_name


def update_excel_with_new_values(target_ws, source_ws, df):    
    max_row = source_ws.max_row
    max_col = source_ws.max_column 

    df = df.toPandas()
    # Copy cell values and styles
    col_name = None
    for row_index in range(0, max_row):
        code = source_ws.cell(row=row_index+1, column=1).value
        for col_inedx in range(0, max_col-1):
            col_name = get_col_name(col_inedx)
            source_cell = source_ws.cell(row=row_index+1, column=col_inedx+1)

            default_cell_format = target_wb.add_format(copy_font(source_cell))

            if str(col_name) not in years or code not in df.code.values:
                # Fall back to formula
                if source_cell.data_type == 'f':
                    cell_format = target_wb.add_format(copy_font(source_cell, blue_text_format=APPLY_BLUE_FONT_IF_MISSING))
                    formula = getattr(source_cell.value, "text", source_cell.value)
                    target_ws.write_formula(row_index, col_inedx, formula, cell_format)  
                else:
                    # Expand formula for years not existent on the original Excel but existent on MEGA
                    if col_name.isdigit() and int(col_name) <= BOOST_LATEST_YEAR and int(col_name) > CCI_LATEST_YEAR:
                        previous_cell = source_ws.cell(row=row_index+1, column=col_inedx)
                        if previous_cell.data_type == 'f':
                            previous_formula = getattr(previous_cell.value, "text", previous_cell.value)
                            current_formula = Translator(previous_formula, origin=previous_cell.coordinate).translate_formula(source_cell.coordinate)
                            target_ws.write_formula(row_index, col_inedx, current_formula, cell_format)
                        else:
                            target_ws.write(row_index, col_inedx, source_cell.value,default_cell_format)
                    else:
                        target_ws.write(row_index, col_inedx, source_cell.value,default_cell_format)
            else:                
                boost_value = df[str(col_name)][df['code'] == code].values[0]
                target_ws.write(row_index, col_inedx, boost_value,default_cell_format)
        set_width(target_ws, max_col)

# COMMAND ----------

# Load an existing workbook

source_wb = load_workbook(SOURCE_FILE_PATH,  data_only=False, read_only=True,)  # Important: data_only=False to get formulas
required_sheets = ["Executed", "Approved"]
for sheet in required_sheets:
    if sheet not in source_wb.sheetnames:
        raise KeyError(f"Required sheet '{sheet}' is missing in the source workbook.")

# pandas.to_excel() cannot write directly to DBFS paths like '/dbfs/...'.
# Workaround: write to a temporary local file, then copy it to the target DBFS location.
with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=True) as tmp:
    temp_path = tmp.name

    # openpyxl was too resource-intensive for writing Excel files with formatting and styles for our use case.
    # Given our cluster limitations, xlsxwriter was the only viable solution for efficient output.
    with pd.ExcelWriter(temp_path, engine='xlsxwriter') as writer:
        raw_data = spark_to_pandas_with_reorder(source_wb["Data_Expenditures"], raw_data, PUBLISH_WITH_BOOST)
        raw_data.to_excel(writer, sheet_name='Data_Expenditures', index=False)

        revenue = spark.table(TARGET_TABLE_REVENUE)
        revenue = spark_to_pandas_with_reorder(source_wb["Data_Revenues"], revenue, PUBLISH_WITH_BOOST)
        revenue.to_excel(writer, sheet_name='Data_Revenues', index=False)

        # Access the xlsxwriter workbook and sheet objects
        target_wb  = writer.book

        # add name manager to the target workbook
        for name, defn in source_wb.defined_names.items():
            if name in ["level", "econ0", "econ_func"]:
                continue
            target_wb.define_name(name, f"={defn.attr_text}")


        # Create an 'Executed' sheet
        target_ws = target_wb.add_worksheet('Executed')  
        source_ws = source_wb['Executed'] 
        update_excel_with_new_values(target_ws, source_ws, executed)

        # Create an 'Approved' sheet
        target_ws = target_wb.add_worksheet('Approved')  
        source_ws = source_wb['Approved'] 
        update_excel_with_new_values(target_ws, source_ws, approved)

    source_wb.close()
    shutil.copy(temp_path, OUTPUT_FILE_PATH)

# COMMAND ----------

coverage_start = raw_data.year.min()
coverage_end = raw_data.year.max()
coverage_country = "Albania"
tagging_sql = f"""
ALTER TABLE {TARGET_TABLE} SET TAGS (
    'name' = 'Albania BOOST platform',
    'ddh_dataset_id' = "0038087",
    'subject' = 'Finance',
    'classification' = 'Public',
    'category' = 'Public Sector',
    'subcategory' = 'Financial Management',
    'frequency' = 'Annually',
    'collections' = 'Financial Management (FM), BOOST - Public Expenditure Database',
    'source' = 'BOOST',
    'domain' = 'Budget',
    'subdomain' = 'Budget & Cost Accounting',
    'excel_link' = '{OUTPUT_FILE_PATH}',
    'license' = 'Creative Commons Attribution-Non Commercial 4.0',
    'topics' = 'Economic Growth, Macroeconomic and Structural Policies, Public Sector Management',
    'coverage_year_start' = '{coverage_start}',
    'coverage_year_end' = '{coverage_end}',
    'coverage_countries' = '{coverage_country}',
    'team_lead' = 'mmastruzzi@worldbank.org',
    'collaborators' = 'icapita@worldbank.org, agirongordillo@worldbank.org, sbhupatiraju@worldbank.org, ysuzuki2@worldbank.org, elysenko@worldbank.org, wlu4@worldbank.org'
);
"""

spark.sql(tagging_sql)

# COMMAND ----------

# Write the missing code descriptions to file for counterpart review

cols_with_labels = ['admin1', 'admin2', 'admin3', 'admin4', 'admin5', 'econ1', 'econ2', 'econ3', 'econ4', 'econ5', 'func1', 'func2', 'program']
df = spark.read.table('prd_mega.boost_intermediate.alb_2023_onward_boost_silver').toPandas()

with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=True) as tmp:
    temp_path = tmp.name

    with pd.ExcelWriter(temp_path, engine='xlsxwriter') as writer:
        for col in cols_with_labels:
            digit_entries = df[col].astype(str)
            digit_entries = digit_entries[digit_entries.str.isdigit()]
            
            if not digit_entries.empty:
                sheet_df = pd.DataFrame({
                    'code': digit_entries.values,
                    'label': [''] * len(digit_entries)
                }).drop_duplicates().sort_values('code')
                sheet_df.to_excel(writer, sheet_name=col, index=False)

        is_missing_lab = df['project_lab'].isnull() | (df['project_lab'].astype(str).str.strip() == '')
        project_codes = df.loc[is_missing_lab, 'project'].dropna().astype(str)
        # handle project_lab seperately since the column will not have the codes. It just only descriptions
        if not project_codes.empty:
            project_lab_df = pd.DataFrame({
                'code': project_codes.values,
                'label': [''] * len(project_codes)
            }).drop_duplicates().sort_values('code')
            project_lab_df.to_excel(writer, sheet_name='project_lab', index=False)
            
    shutil.copy(temp_path, OUTPUT_MISSING_DESC_FILE_PATH)
