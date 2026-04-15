"""Flag columns in Approved/Executed where values were embedded (hard-coded)
instead of computed by formula. These won't be reproduced by the pipeline
and need expert review.

Columns whose formulas evaluate to nothing (no underlying data for that year)
are tagged NO_DATA and excluded from the suspicious classifications.

Outputs:
  - hardcoded_columns.csv : per (sheet, column) summary — formula vs literal counts
  - hardcoded_cells.csv   : every literal-value cell (excluding header row) for inspection
"""
from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

MISSING_MARKERS = {"..", ".", "n/a", "N/A", "-", "–", "—", ""}


def is_missing(v) -> bool:
    """Treat blanks and conventional 'no data' markers as missing — not overrides."""
    if v is None:
        return True
    if isinstance(v, str) and v.strip() in MISSING_MARKERS:
        return True
    return False

XLSX = Path("/Users/ysuzuki2/Developer/mega-boost/temp/Zimbabwe BOOST.xlsx")
ROOT = Path("/Users/ysuzuki2/Developer/mega-boost/Zimbabwe/_analysis")
OUT = ROOT / "data"
OUT.mkdir(parents=True, exist_ok=True)
TARGET_SHEETS = ("Approved", "Executed")
HEADER_ROWS = 1  # treat row 1 as header (labels)


def main():
    wb = load_workbook(XLSX, read_only=True, data_only=False)
    wb_v = load_workbook(XLSX, read_only=True, data_only=True)  # cached values for NO_DATA detection

    col_summary = []          # rows for hardcoded_columns.csv
    cell_rows = []            # rows for hardcoded_cells.csv

    for sheet in TARGET_SHEETS:
        if sheet not in wb.sheetnames:
            print(f"skip: {sheet} not found")
            continue
        ws = wb[sheet]
        ws_v = wb_v[sheet]

        # per-column tallies
        n_formula = defaultdict(int)
        n_literal = defaultdict(int)        # numeric overrides only — these are real issues
        n_label   = defaultdict(int)        # string literals — labels/tags, not flagged
        n_blank   = defaultdict(int)
        # cached-value tallies — used to detect NO_DATA columns
        # (formula present but evaluates to None/0/empty for the entire column)
        n_formula_empty_value = defaultdict(int)
        headers   = {}
        sample_lit = defaultdict(list)  # col -> [(coord, value), ...] up to 5

        v_grid = list(ws_v.iter_rows(values_only=True))

        for r_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            for c_idx, cell in enumerate(row, start=1):
                v = cell.value
                if r_idx <= HEADER_ROWS:
                    if v is not None:
                        headers[c_idx] = str(v)
                    continue
                if is_missing(v):
                    n_blank[c_idx] += 1
                    continue
                # array formulas count as formulas, not literals
                is_formula = (isinstance(v, str) and v.startswith("=")) or isinstance(v, ArrayFormula)
                if is_formula:
                    n_formula[c_idx] += 1
                    cached = None
                    if r_idx - 1 < len(v_grid) and c_idx - 1 < len(v_grid[r_idx - 1]):
                        cached = v_grid[r_idx - 1][c_idx - 1]
                    if is_missing(cached) or cached == 0:
                        n_formula_empty_value[c_idx] += 1
                else:
                    # only numeric literals count as "overrides" — strings are labels/tags
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        n_literal[c_idx] += 1
                        if len(sample_lit[c_idx]) < 5:
                            sample_lit[c_idx].append((cell.coordinate, v))
                        cell_rows.append({
                            "sheet": sheet,
                            "cell": cell.coordinate,
                            "row": r_idx,
                            "col_letter": get_column_letter(c_idx),
                            "col_idx": c_idx,
                            "header": headers.get(c_idx, ""),
                            "value": v,
                            "type": type(v).__name__,
                        })
                    else:
                        n_label[c_idx] += 1

        # Override headers with cached values so year columns render as 2018, 2019, ...
        # instead of the raw formula "=N1+1".
        if v_grid:
            for c_idx, hv in enumerate(v_grid[0], start=1):
                if hv is None or hv == "":
                    continue
                if isinstance(hv, float) and hv.is_integer():
                    headers[c_idx] = str(int(hv))
                else:
                    headers[c_idx] = str(hv)

        # build per-column summary
        all_cols = set(n_formula) | set(n_literal) | set(n_blank) | set(headers)
        for c_idx in sorted(all_cols):
            f = n_formula[c_idx]
            l = n_literal[c_idx]
            b = n_blank[c_idx]
            f_empty = n_formula_empty_value[c_idx]
            total_nonblank = f + l
            classification = "empty"
            if total_nonblank:
                if f == 0 and l > 0:
                    classification = "ALL_LITERAL"        # whole column hand-entered
                elif f > 0 and l == 0:
                    classification = "all_formula"
                elif f > 0 and l > 0:
                    classification = "MIXED_LITERAL_FORMULA"  # most suspicious
                # demote if the formulas in this column produce nothing —
                # that's a "no data for this year" column, not an issue.
                if f > 0 and f_empty == f:
                    classification = "NO_DATA"
            col_summary.append({
                "sheet": sheet,
                "col_letter": get_column_letter(c_idx),
                "col_idx": c_idx,
                "header": headers.get(c_idx, ""),
                "n_formula": f,
                "n_formula_empty_value": f_empty,
                "n_literal_numeric": l,
                "n_label_string": n_label[c_idx],
                "n_blank": b,
                "literal_pct_of_filled": round(100 * l / total_nonblank, 1) if total_nonblank else 0,
                "classification": classification,
                "literal_samples": "; ".join(f"{c}={v!r}" for c, v in sample_lit[c_idx]),
            })

    with (OUT / "hardcoded_columns.csv").open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(col_summary[0].keys()))
        w.writeheader()
        w.writerows(col_summary)

    with (OUT / "hardcoded_cells.csv").open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["sheet","cell","row","col_letter","col_idx","header","value","type"])
        w.writeheader()
        w.writerows(cell_rows)

    # console summary — focus on suspicious classifications
    print(f"Wrote hardcoded_columns.csv ({len(col_summary)} rows)")
    print(f"Wrote hardcoded_cells.csv   ({len(cell_rows)} rows)\n")
    print("Columns with hand-entered NUMERIC values (real overrides):\n")
    print(f"  {'sheet':<10} {'column header':<35} {'n_formula':>9} {'n_num':>6} class")
    for r in col_summary:
        if r["classification"] in ("ALL_LITERAL", "MIXED_LITERAL_FORMULA"):
            label = (r["header"] or r["col_letter"])[:35]
            print(f"  {r['sheet']:<10} {label:<35} {r['n_formula']:>9} {r['n_literal_numeric']:>6} {r['classification']}")


if __name__ == "__main__":
    main()
