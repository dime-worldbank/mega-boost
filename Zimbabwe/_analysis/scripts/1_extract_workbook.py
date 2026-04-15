"""Phase 1: Excel archaeology for Zimbabwe BOOST workbook.

Produces:
  - sheet_inventory.csv   : every sheet + dims + role guess
  - formula_map.csv       : every formula cell in Approved/Executed with refs + archetype
  - data_dictionary.csv   : parsed dictionary tab(s)
  - FINDINGS.md           : human-readable summary

Run locally (openpyxl only, no Spark).
"""
from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

XLSX = Path("/Users/ysuzuki2/Developer/mega-boost/temp/Zimbabwe BOOST.xlsx")
ROOT = Path("/Users/ysuzuki2/Developer/mega-boost/Zimbabwe/_analysis")
OUT = ROOT / "data"
OUT.mkdir(parents=True, exist_ok=True)

FORMULA_SHEETS = {"Approved", "Executed"}

# crude archetype classifier
ARCHETYPES = [
    ("SUMIFS",   re.compile(r"\bSUMIFS\s*\(", re.I)),
    ("SUMIF",    re.compile(r"\bSUMIF\s*\(", re.I)),
    ("SUMPRODUCT", re.compile(r"\bSUMPRODUCT\s*\(", re.I)),
    ("XLOOKUP",  re.compile(r"\bXLOOKUP\s*\(", re.I)),
    ("VLOOKUP",  re.compile(r"\bVLOOKUP\s*\(", re.I)),
    ("INDEX_MATCH", re.compile(r"\bINDEX\s*\(.*\bMATCH\s*\(", re.I | re.S)),
    ("IFS",      re.compile(r"\bIFS\s*\(", re.I)),
    ("IF",       re.compile(r"\bIF\s*\(", re.I)),
    ("SUM",      re.compile(r"\bSUM\s*\(", re.I)),
    ("DIRECT_REF", re.compile(r"^=\s*'?[^=!()+\-*/,]+!", re.I)),
    ("ARITH",    re.compile(r"^=[^A-Za-z]*$")),
]

# sheet!range or 'sheet name'!range
SHEETREF_RE = re.compile(r"(?:'([^']+)'|([A-Za-z_][\w\.]*))!\$?[A-Z]+\$?\d*(?::\$?[A-Z]+\$?\d*)?")
# bare identifier (defined name) — word not followed by ( and not a cell ref like A1
IDENT_RE = re.compile(r"(?<![A-Za-z0-9_!'])([A-Za-z_][A-Za-z0-9_]*)(?!\s*\()")
CELLREF_RE = re.compile(r"^\$?[A-Z]{1,3}\$?\d+$")
EXCEL_FUNCS = {
    "IF","IFS","AND","OR","NOT","SUM","SUMIF","SUMIFS","SUMPRODUCT","COUNT","COUNTA",
    "COUNTIF","COUNTIFS","AVERAGE","MIN","MAX","ROUND","ROUNDUP","ROUNDDOWN","ABS",
    "VLOOKUP","HLOOKUP","XLOOKUP","INDEX","MATCH","LOOKUP","OFFSET","INDIRECT",
    "ISBLANK","ISNUMBER","ISERROR","ISERR","IFERROR","IFNA","TRUE","FALSE",
    "LEFT","RIGHT","MID","LEN","TRIM","UPPER","LOWER","CONCATENATE","TEXT","VALUE",
    "YEAR","MONTH","DAY","DATE","TODAY","NOW","ROW","COLUMN","ROWS","COLUMNS",
    "N","T","NA","ERROR.TYPE","REPT","SUBSTITUTE","FIND","SEARCH",
}


def classify(formula: str) -> str:
    if not formula:
        return ""
    for name, pat in ARCHETYPES:
        if pat.search(formula):
            return name
    return "OTHER"


def extract_refs(formula: str, defined_names: dict[str, str]) -> tuple[list[str], list[str]]:
    """Return (sheet_refs, defined_name_refs)."""
    sheets = set()
    for m in SHEETREF_RE.finditer(formula):
        sheet = m.group(1) or m.group(2)
        sheets.add(sheet.strip())
    # strip out literal strings before hunting identifiers
    stripped = re.sub(r'"[^"]*"', "", formula)
    names = set()
    for m in IDENT_RE.finditer(stripped):
        tok = m.group(1)
        if tok.upper() in EXCEL_FUNCS:
            continue
        if CELLREF_RE.match(tok):
            continue
        if tok in defined_names:
            names.add(tok)
            # also add the target sheet
            target = defined_names[tok]
            sm = SHEETREF_RE.search(target)
            if sm:
                sheets.add((sm.group(1) or sm.group(2)).strip())
    return sorted(sheets), sorted(names)


def guess_role(name: str, max_row: int, max_col: int, has_formulas: bool) -> str:
    lname = name.lower()
    if any(k in lname for k in ("approved", "executed", "revised")):
        return "formula_output"
    if any(k in lname for k in ("dict", "label", "lookup", "mapping", "code", "cofog", "econ_", "func_")):
        return "dictionary"
    if any(k in lname for k in ("raw", "data", "microdata", "detail", "source")):
        return "raw"
    if has_formulas and max_row * max_col < 5000:
        return "summary/formula"
    if max_row > 500:
        return "raw"
    return "unknown"


def main():
    print(f"Loading {XLSX} (formulas pass)…")
    wb_f = load_workbook(XLSX, read_only=True, data_only=False)
    print(f"Loading {XLSX} (cached values pass)…")
    wb_v = load_workbook(XLSX, read_only=True, data_only=True)

    # defined names (named ranges) — critical because Zimbabwe formulas use them
    defined_names: dict[str, str] = {}
    for dn_name in wb_f.defined_names:
        try:
            defined_names[dn_name] = wb_f.defined_names[dn_name].value or ""
        except Exception:
            pass
    with (OUT / "defined_names.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["name", "target"])
        for k, v in sorted(defined_names.items()):
            w.writerow([k, v])
    print(f"Extracted {len(defined_names)} defined names → defined_names.csv")

    # ---- sheet_inventory ----
    inventory_rows = []
    per_sheet_formula_count = {}
    for name in wb_f.sheetnames:
        ws = wb_f[name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        formula_count = 0
        # sample-scan: formula presence (early exit once we know)
        if max_row and max_col:
            cap_rows = min(max_row, 2000)
            for row in ws.iter_rows(min_row=1, max_row=cap_rows, values_only=False):
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        formula_count += 1
                        if formula_count > 5 and name not in FORMULA_SHEETS:
                            break
                if formula_count > 5 and name not in FORMULA_SHEETS:
                    break
        per_sheet_formula_count[name] = formula_count
        role = guess_role(name, max_row, max_col, formula_count > 0)
        inventory_rows.append({
            "sheet": name,
            "max_row": max_row,
            "max_col": max_col,
            "formula_cells_sampled": formula_count,
            "role_guess": role,
        })
        print(f"  {name:40s} rows={max_row:<7} cols={max_col:<4} formulas~{formula_count:<5} role={role}")

    with (OUT / "sheet_inventory.csv").open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(inventory_rows[0].keys()))
        w.writeheader()
        w.writerows(inventory_rows)

    # ---- formula_map for Approved & Executed ----
    target_sheets = [n for n in wb_f.sheetnames if n.strip() in FORMULA_SHEETS]
    if not target_sheets:
        # fuzzy match
        target_sheets = [n for n in wb_f.sheetnames if any(k in n.lower() for k in ("approv", "execut"))]
    print(f"\nFormula-target sheets: {target_sheets}")

    archetype_counter = Counter()
    ref_counter = Counter()
    name_counter = Counter()
    formula_rows = []
    for name in target_sheets:
        ws_f = wb_f[name]
        ws_v = wb_v[name]
        v_iter = ws_v.iter_rows(values_only=True)
        v_grid = list(v_iter)
        for r_idx, row in enumerate(ws_f.iter_rows(values_only=False), start=1):
            for c_idx, cell in enumerate(row, start=1):
                val = cell.value
                if not (isinstance(val, str) and val.startswith("=")):
                    continue
                archetype = classify(val)
                refs, names = extract_refs(val, defined_names)
                # reclassify OTHER when only defined-name math / direct lookups
                if archetype == "OTHER" and names:
                    archetype = "NAMED_REF"
                archetype_counter[archetype] += 1
                for r in refs:
                    ref_counter[r] += 1
                for n in names:
                    name_counter[n] += 1
                cached = None
                if r_idx - 1 < len(v_grid) and c_idx - 1 < len(v_grid[r_idx - 1]):
                    cached = v_grid[r_idx - 1][c_idx - 1]
                formula_rows.append({
                    "sheet": name,
                    "cell": cell.coordinate,
                    "row": r_idx,
                    "col": c_idx,
                    "archetype": archetype,
                    "referenced_sheets": "|".join(refs),
                    "named_ranges": "|".join(names),
                    "cached_value": cached,
                    "formula": val,
                })

    with (OUT / "formula_map.csv").open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=[
            "sheet", "cell", "row", "col", "archetype",
            "referenced_sheets", "named_ranges", "cached_value", "formula",
        ])
        w.writeheader()
        w.writerows(formula_rows)

    print(f"\nWrote {len(formula_rows)} formulas to formula_map.csv")
    print("Archetypes:", archetype_counter.most_common())
    print("Top referenced sheets:", ref_counter.most_common(15))
    print("Top named ranges:", name_counter.most_common(20))

    # ---- data_dictionary ----
    dict_sheets = [r["sheet"] for r in inventory_rows if r["role_guess"] == "dictionary"]
    dd_rows = []
    for name in dict_sheets:
        ws = wb_v[name]
        header = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c) if c is not None else "" for c in row]
                continue
            if all(c is None for c in row):
                continue
            record = {"_sheet": name}
            for h, v in zip(header, row):
                record[h or "_blank"] = v
            dd_rows.append(record)
    if dd_rows:
        all_keys = sorted({k for r in dd_rows for k in r.keys()})
        with (OUT / "data_dictionary.csv").open("w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=all_keys)
            w.writeheader()
            w.writerows(dd_rows)
        print(f"Wrote data_dictionary.csv with {len(dd_rows)} rows from {dict_sheets}")
    else:
        print("No dictionary sheets auto-detected; review sheet_inventory.csv and re-classify.")

    # ---- FINDINGS.md ----
    lines = []
    lines.append("# Zimbabwe BOOST — Phase 1 Findings\n")
    lines.append(f"Source: `{XLSX.name}` (23MB)\n")
    lines.append("## Sheets\n")
    lines.append("| Sheet | Rows | Cols | Formulas (sampled) | Role guess |")
    lines.append("|---|---:|---:|---:|---|")
    for r in inventory_rows:
        lines.append(f"| {r['sheet']} | {r['max_row']} | {r['max_col']} | {r['formula_cells_sampled']} | {r['role_guess']} |")
    lines.append("\n## Formula archetypes (Approved + Executed)\n")
    total_f = sum(archetype_counter.values()) or 1
    for a, n in archetype_counter.most_common():
        lines.append(f"- **{a}**: {n} ({n*100//total_f}%)")
    lines.append("\n## Top referenced sheets from formulas\n")
    for s, n in ref_counter.most_common(15):
        lines.append(f"- `{s}`: {n} refs")
    lines.append("\n## Top named ranges used in formulas\n")
    lines.append("| Name | Target | Usage count |")
    lines.append("|---|---|---:|")
    for nm, cnt in name_counter.most_common(25):
        lines.append(f"| `{nm}` | `{defined_names.get(nm, '?')}` | {cnt} |")
    lines.append("\n## Next steps\n")
    lines.append("- Review `sheet_inventory.csv` and confirm role guesses (especially `unknown`).")
    lines.append("- Spot-check `formula_map.csv` — confirm archetype classification is accurate.")
    lines.append("- Inspect `data_dictionary.csv`; if empty, identify dict sheets manually and rerun.")
    lines.append("- Phase 2: hand-translate each archetype group into PySpark in `ZWE_transform_load_raw_dlt.py`.")
    reports_dir = ROOT / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    provenance = [
        "> Generated by `scripts/1_extract_workbook.py` from the source workbook.",
        "> Rerun with `python3 scripts/1_extract_workbook.py`.\n",
    ]
    (reports_dir / "workbook_inventory.md").write_text("\n".join(provenance + lines))
    print(f"\nWrote reports/workbook_inventory.md")


if __name__ == "__main__":
    main()
