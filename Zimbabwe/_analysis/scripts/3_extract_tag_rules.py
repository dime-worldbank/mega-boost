"""Extract a tag-rules table from Approved/Executed.

Each row in those sheets is a BOOST classification tag. Its SUMIFS criteria
define the filter on the Expenditure raw data. We parse those criteria
verbatim — no interpretation — into a structured table that the silver layer
will iterate over.

Output:
  tag_rules.csv  one row per Approved-sheet (or Executed-sheet) row, with:
    sheet, row, code, category, fiscal_year, region, income, basis, source,
    rigidity, formulas_text, coverage_text, sample_formula, sample_year,
    measure, criteria_json

The `criteria_json` is a JSON list of {field, op, value} dicts, e.g.:
  [{"field":"Exclude","op":"<>","value":"y"},
   {"field":"econ3","op":"=","value":"606"}]

Other countries hand-code their classification logic in PySpark; Zimbabwe is
unusual in that the workbook itself encodes the classification as SUMIFS
criteria. This script lifts that encoding so the pipeline can apply it
literally.
"""
from __future__ import annotations

import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

XLSX = Path("/Users/ysuzuki2/Developer/mega-boost/temp/Zimbabwe BOOST.xlsx")
ROOT = Path("/Users/ysuzuki2/Developer/mega-boost/Zimbabwe/_analysis")
OUT = ROOT / "data"
OUT.mkdir(parents=True, exist_ok=True)

TARGET_SHEETS = ("Approved", "Executed")

# Column map for Approved/Executed rows (1-indexed cells).
COLS = {
    "code":         "A",
    "category":     "B",
    "formulas":     "AE",
    "coverage":     "AF",
    "fiscal_year":  "AG",
    "region":       "AH",
    "income":       "AI",
    "basis":        "AJ",
    "source":       "AK",
    "rigidity":     "AL",
}

# Year header columns: scan row 1 cached values; treat 4-digit numbers as years.
YEAR_HEADER_RE = re.compile(r"^\s*(19|20)\d{2}\s*$")


def split_sumifs_args(inner: str) -> list[str]:
    """Split the inside of a SUMIFS(...) call by top-level commas, respecting
    quoted strings and nested parens."""
    out, buf, depth, in_str = [], [], 0, False
    for ch in inner:
        if ch == '"' and (not buf or buf[-1] != "\\"):
            in_str = not in_str
            buf.append(ch)
            continue
        if in_str:
            buf.append(ch)
            continue
        if ch == "(":
            depth += 1
            buf.append(ch)
        elif ch == ")":
            depth -= 1
            buf.append(ch)
        elif ch == "," and depth == 0:
            out.append("".join(buf).strip())
            buf = []
        else:
            buf.append(ch)
    if buf:
        out.append("".join(buf).strip())
    return out


def parse_criterion(value: str) -> tuple[str, str]:
    """Return (op, literal). Excel criteria can be:
      "<>y", ">=10", "606", "<2024", year-cell-ref like O$1.
    Cell refs are returned as op="=year" so the caller can substitute the year.
    """
    v = value.strip()
    # quoted string
    if v.startswith('"') and v.endswith('"'):
        v = v[1:-1]
        for op in ("<>", "<=", ">=", "<", ">", "="):
            if v.startswith(op):
                return op, v[len(op):]
        return "=", v
    # cell reference (year header lookup) — column letter + $row+digits
    if re.match(r"^\$?[A-Z]{1,3}\$?\d+$", v):
        return "=year", v
    # numeric / unquoted literal
    return "=", v


# Match SUMIFS(...) including nested parens.
SUMIFS_RE = re.compile(r"SUMIFS\s*\(", re.I)


def extract_sumifs_blocks(formula: str) -> list[str]:
    """Return every top-level SUMIFS(...) inner-arg-string in the formula."""
    blocks = []
    i = 0
    while True:
        m = SUMIFS_RE.search(formula, i)
        if not m:
            break
        depth = 1
        j = m.end()
        while j < len(formula) and depth:
            if formula[j] == "(":
                depth += 1
            elif formula[j] == ")":
                depth -= 1
            j += 1
        blocks.append(formula[m.end(): j - 1])
        i = j
    return blocks


def parse_criteria(formula: str) -> tuple[str, list[dict]]:
    """Return (measure, criteria) parsed from the FIRST SUMIFS in a formula.
    measure is the named range / sheet ref summed (e.g., 'approved', 'executed').
    """
    blocks = extract_sumifs_blocks(formula)
    if not blocks:
        return "", []
    args = split_sumifs_args(blocks[0])
    if len(args) < 3:
        return args[0] if args else "", []
    measure = args[0]
    criteria = []
    for k in range(1, len(args) - 1, 2):
        field = args[k].strip()
        raw_value = args[k + 1].strip()
        op, lit = parse_criterion(raw_value)
        criteria.append({"field": field, "op": op, "value": lit})
    return measure, criteria


def main():
    wb_f = load_workbook(XLSX, read_only=True, data_only=False)
    wb_v = load_workbook(XLSX, read_only=True, data_only=True)

    out_rows = []
    for sheet in TARGET_SHEETS:
        if sheet not in wb_f.sheetnames:
            continue
        ws_f = wb_f[sheet]
        ws_v = wb_v[sheet]
        v_grid = list(ws_v.iter_rows(values_only=True))
        f_grid = list(ws_f.iter_rows(values_only=False))

        # Identify year-header columns from row 1 cached values.
        year_cols: list[tuple[int, int]] = []  # (col_idx, year)
        if v_grid:
            for c_idx, hv in enumerate(v_grid[0], start=1):
                if hv is None:
                    continue
                hv_str = str(int(hv)) if isinstance(hv, float) and hv.is_integer() else str(hv).strip()
                if YEAR_HEADER_RE.match(hv_str):
                    year_cols.append((c_idx, int(hv_str)))

        col_idx = {k: column_index_from_string(v) for k, v in COLS.items()}

        for r_idx, row in enumerate(f_grid, start=1):
            if r_idx == 1:
                continue
            code_cell = ws_v.cell(r_idx, col_idx["code"]).value
            if code_cell is None or str(code_cell).strip() == "":
                continue

            # Scan year columns and classify the row.
            # SUMIFS rows define raw-data tag rules; plain-SUM rows are rollups
            # of other Approved/Executed rows; HEADER rows have no formula.
            sample_formula = ""
            sample_year = None
            row_type = "HEADER"
            rollup_refs = ""
            for c_idx, year in year_cols:
                cell = ws_f.cell(r_idx, c_idx).value
                if not isinstance(cell, str) or not cell.startswith("="):
                    continue
                if "SUMIFS" in cell.upper():
                    sample_formula = cell
                    sample_year = year
                    row_type = "TAG"
                    break
                if row_type == "HEADER":
                    sample_formula = cell
                    sample_year = year
                    row_type = "ROLLUP"
                    rollup_refs = cell  # keep first non-SUMIFS formula for traceability

            measure, criteria = parse_criteria(sample_formula) if row_type == "TAG" else ("", [])

            out_rows.append({
                "sheet":         sheet,
                "row":           r_idx,
                "row_type":      row_type,
                "code":          code_cell,
                "category":      ws_v.cell(r_idx, col_idx["category"]).value,
                "fiscal_year":   ws_v.cell(r_idx, col_idx["fiscal_year"]).value,
                "region":        ws_v.cell(r_idx, col_idx["region"]).value,
                "income":        ws_v.cell(r_idx, col_idx["income"]).value,
                "basis":         ws_v.cell(r_idx, col_idx["basis"]).value,
                "source":        ws_v.cell(r_idx, col_idx["source"]).value,
                "rigidity":      ws_v.cell(r_idx, col_idx["rigidity"]).value,
                "formulas_text": ws_v.cell(r_idx, col_idx["formulas"]).value,
                "coverage_text": ws_v.cell(r_idx, col_idx["coverage"]).value,
                "sample_formula": sample_formula,
                "sample_year":   sample_year,
                "measure":       measure,
                "criteria_json": json.dumps(criteria, ensure_ascii=False),
            })

    headers = list(out_rows[0].keys()) if out_rows else []
    with (OUT / "tag_rules.csv").open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        w.writerows(out_rows)

    # Console summary
    print(f"Wrote {len(out_rows)} tag rules → tag_rules.csv")
    by_sheet_type = {}
    by_field_count = {}
    fields_used = {}
    measures = {}
    for r in out_rows:
        key = (r["sheet"], r["row_type"])
        by_sheet_type[key] = by_sheet_type.get(key, 0) + 1
        if r["row_type"] != "TAG":
            continue
        crits = json.loads(r["criteria_json"])
        by_field_count[len(crits)] = by_field_count.get(len(crits), 0) + 1
        for c in crits:
            fields_used[c["field"]] = fields_used.get(c["field"], 0) + 1
        measures[r["measure"]] = measures.get(r["measure"], 0) + 1
    print("  rows by sheet × type:")
    for k, v in sorted(by_sheet_type.items()):
        print(f"    {k[0]:<10} {k[1]:<8} {v}")
    print(f"  TAG rows criterion-count distribution: {dict(sorted(by_field_count.items()))}")
    print(f"  TAG fields used in criteria: {sorted(fields_used.items(), key=lambda x: -x[1])}")
    print(f"  TAG measures summed: {measures}")


if __name__ == "__main__":
    main()
