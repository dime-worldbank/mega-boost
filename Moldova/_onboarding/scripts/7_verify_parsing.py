"""Phase 7 — local formula-parsing verification.

For every TAG rule in `data/tag_rules.csv`, apply its parsed `criteria_json`
(OR across SUMIFS blocks, AND within a block) to the matching year-range
raw sheet and compare the per-year sum against Excel's cached cell value
in the Approved / Executed sheet. Any mismatch points at:

  - a parser bug in `3_extract_tag_rules.py` (e.g. unhandled operator,
    array-constant mis-split, localized value typo);
  - a formula feature we don't yet model (IF-wrapping, cell subtraction
    like `…-C19`, INDIRECT/OFFSET, externals);
  - a hand-entered override (see `reports/ISSUES.md §1`) — legitimately
    unreproducible from SUMIFS alone.

Output:
  - `reports/parsing_verification.md` — per-rule pass/fail summary with
    the worst-offender cells for manual triage
  - `data/parsing_verification.csv` — one row per (rule, year) check for
    drill-down

Runs locally against the source workbook (no Spark). Expects
`temp/Moldova BOOST.xlsx` to be present.
"""
from __future__ import annotations

import csv
import json
import math
import re
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent          # Moldova/_onboarding
COUNTRY_DIR = ROOT.parent                              # Moldova/
XLSX = COUNTRY_DIR.parent / "temp" / "Moldova BOOST.xlsx"
DATA = ROOT / "data"
REPORTS = ROOT / "reports"
REPORTS.mkdir(parents=True, exist_ok=True)
# Driver CSV (tag_rules.csv) lives at COUNTRY_DIR for DLT pipeline consumption.
TAG_RULES = COUNTRY_DIR / "tag_rules.csv"

# Named-range → raw-column mapping per year-range sheet. Must match the
# silver-layer maps in MDA_transform_load_raw_dlt.py.
MAP_BASE = {  # 2006-15
    "year": "year", "admin1": "admin1",
    "func1": "func1", "func2": "func2",
    "econ1": "econ1", "econ2": "econ2",
    "exp_type": "exp_type", "transfer": "transfer",
    "approved": "approved", "executed": "executed", "adjusted": "adjusted",
}
MAP_16 = {  # 2016-19
    "year_16": "year",
    "admin1_16": "admin1", "admin2_16": "admin2",
    "func1_16": "func1", "func2_16": "func2", "func3_16": "func3",
    "econ0_16": "econ0",
    "econ1_16": "econ1", "econ2_16": "econ2", "econ3_16": "econ3",
    "econ4_16": "econ4", "econ5_16": "econ5", "econ6_16": "econ6",
    "exp_type_16": "exp_type", "transfer_16": "transfer",
    "fin_source1_16": "fin_source1",
    "approved_16": "approved", "revised_16": "revised",
    "executed_16": "executed",
    "program1_16": "program1", "program2_16": "program2",
    "activity_16": "activity",
}
MAP_20 = {  # 2020-24
    "year_20": "year",
    "admin1_20": "admin1", "admin2_20": "admin2",
    "func1_20": "func1", "func2_20": "func2", "func3_20": "func3",
    "econ0_20": "econ0",
    "econ1_20": "econ1", "econ2_20": "econ2", "econ3_20": "econ3",
    "econ4_20": "econ4", "econ5_20": "econ5", "econ6_20": "econ6",
    "exp_type_20": "exp_type", "transfer_20": "transfer",
    "fin_source1_20": "fin_source1",
    "approved_20": "approved", "adjusted_20": "adjusted",
    "executed_20": "executed",
}

RANGE_CONFIG = {
    "base": {"sheet": "2006-15", "mapping": MAP_BASE},
    "16":   {"sheet": "2016-19", "mapping": MAP_16},
    "20":   {"sheet": "2020-24", "mapping": MAP_20},
}

MATCH_EPSILON = 0.5  # sub-unit currency tolerance (Excel rounds internally)


def classify_measure(measure: str) -> str | None:
    if measure in ("approved", "executed"):
        return "base"
    if measure.endswith("_16"):
        return "16"
    if measure.endswith("_20"):
        return "20"
    return None  # Raw2 and other non-standard measures are skipped here


def measure_column(measure: str, mapping: dict[str, str]) -> str | None:
    """Resolve a rule's measure to the raw-sheet column to sum."""
    return mapping.get(measure)


# ---------- criterion evaluation (pandas) ----------

def _cmp(series: pd.Series, op: str, value: str) -> pd.Series:
    if "*" in value:
        pattern = "^" + re.escape(value).replace(r"\*", ".*") + "$"
        m = series.astype(str).str.match(pattern, case=False, na=False)
        return ~m if op == "<>" else m
    if op in ("=", "<>"):
        eq = series.astype(str).str.lower() == str(value).lower()
        return ~eq if op == "<>" else eq
    s = pd.to_numeric(series, errors="coerce")
    v = float(value)
    return {"<": s < v, "<=": s <= v, ">": s > v, ">=": s >= v}[op]


def rule_mask(df: pd.DataFrame, criteria_groups: list[list[dict]],
              mapping: dict[str, str]) -> pd.Series:
    """OR across SUMIFS blocks; AND within a block. A block referencing any
    field not in `mapping` contributes False (the pipeline cannot evaluate
    it — same conservative choice as the silver layer)."""
    if not criteria_groups:
        return pd.Series(False, index=df.index)
    combined = pd.Series(False, index=df.index)
    for branch in criteria_groups:
        branch_mask = pd.Series(True, index=df.index)
        evaluable = False
        for c in branch:
            if c["op"] == "=year":
                continue
            col = mapping.get(c["field"])
            if col is None or col not in df.columns:
                branch_mask = pd.Series(False, index=df.index)
                break
            branch_mask &= _cmp(df[col], c["op"], c["value"])
            evaluable = True
        if evaluable:
            combined |= branch_mask
    return combined


# ---------- workbook access ----------

def load_cached_cells() -> dict[tuple[str, int], dict[int, float | str | None]]:
    """Return {(sheet, row): {year: cached_value}} for Approved/Executed."""
    wb_v = load_workbook(XLSX, read_only=True, data_only=True)
    out: dict[tuple[str, int], dict[int, object]] = {}
    for sheet_name in ("Approved", "Executed"):
        ws = wb_v[sheet_name]
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))]
        year_cols = {int(h): i for i, h in enumerate(hdr, start=1)
                     if isinstance(h, (int, float)) and 2000 <= h <= 2030}
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if r_idx == 1:
                continue
            per_year = {}
            for year, c_idx in year_cols.items():
                if c_idx - 1 < len(row):
                    per_year[year] = row[c_idx - 1]
            out[(sheet_name, r_idx)] = per_year
    return out


def _to_num(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) if not (isinstance(v, float) and math.isnan(v)) else None
    if isinstance(v, str):
        s = v.strip()
        if s in ("", "..", ".", "-", "—", "N/A", "n/a"):
            return None
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    return None


# ---------- verification ----------

def verify():
    print("Loading cached cell values (Approved / Executed)…")
    cached = load_cached_cells()

    print("Loading raw sheets…")
    raw = {}
    for rk, cfg in RANGE_CONFIG.items():
        print(f"  {cfg['sheet']}…")
        raw[rk] = pd.read_excel(XLSX, sheet_name=cfg["sheet"])
        print(f"    {len(raw[rk]):,} rows × {len(raw[rk].columns)} cols")

    with TAG_RULES.open() as f:
        rules = [r for r in csv.DictReader(f) if r["row_type"] == "TAG"]
    print(f"Verifying {len(rules)} TAG rules…\n")

    detail_rows: list[dict] = []
    for r in rules:
        rk = classify_measure(r["measure"])
        if rk is None:
            # Raw2 / other non-standard measures — flagged as such and skipped.
            detail_rows.append({
                "sheet": r["sheet"], "row": int(r["row"]), "code": r["code"],
                "measure": r["measure"], "range": "N/A",
                "year": None, "excel": None, "pipeline": None,
                "status": "SKIP_UNSUPPORTED_MEASURE", "diff": None,
            })
            continue
        cfg = RANGE_CONFIG[rk]
        df = raw[rk]
        mapping = cfg["mapping"]
        measure_col = measure_column(r["measure"], mapping)
        groups = json.loads(r["criteria_json"])

        years = [int(y) for y in r["years_covered"].split(",") if y]
        mask_base = rule_mask(df, groups, mapping)
        cached_row = cached.get((r["sheet"], int(r["row"])), {})

        for year in years:
            year_mask = mask_base & (df["year"] == year)
            if measure_col and measure_col in df.columns:
                pipeline = pd.to_numeric(df.loc[year_mask, measure_col], errors="coerce").sum()
            else:
                pipeline = None
            excel = _to_num(cached_row.get(year))

            if excel is None and pipeline is None:
                status = "BOTH_EMPTY"
                diff = None
            elif excel is None:
                status = "EXCEL_EMPTY" if (pipeline is None or abs(pipeline) < MATCH_EPSILON) else "EXCEL_EMPTY_PIPELINE_NONZERO"
                diff = pipeline if pipeline is not None else None
            elif pipeline is None:
                status = "PIPELINE_NONE"
                diff = -excel
            else:
                diff = pipeline - excel
                status = "MATCH" if abs(diff) <= MATCH_EPSILON else "MISMATCH"
            detail_rows.append({
                "sheet": r["sheet"], "row": int(r["row"]), "code": r["code"],
                "measure": r["measure"], "range": rk,
                "year": year, "excel": excel, "pipeline": pipeline,
                "status": status,
                "diff": diff,
                "n_sumifs": int(r.get("n_sumifs") or 0),
                "sample_formula": r.get("sample_formula", ""),
            })

    # Write detail CSV
    with (DATA / "parsing_verification.csv").open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(detail_rows[0].keys()))
        w.writeheader()
        w.writerows(detail_rows)

    # ---- summary ----
    by_status: dict[str, int] = defaultdict(int)
    for d in detail_rows:
        by_status[d["status"]] += 1
    mismatches = [d for d in detail_rows if d["status"] == "MISMATCH"]
    # Group mismatches by (code, range) to produce a per-rule view
    by_code_range: dict[tuple[str, str, str], list[dict]] = defaultdict(list)
    for d in mismatches:
        by_code_range[(d["code"], d["sheet"], d["range"])].append(d)

    print("=== Summary ===")
    for k in ("MATCH", "MISMATCH", "BOTH_EMPTY", "EXCEL_EMPTY",
             "EXCEL_EMPTY_PIPELINE_NONZERO", "PIPELINE_NONE",
             "SKIP_UNSUPPORTED_MEASURE"):
        if by_status.get(k):
            print(f"  {k:<34} {by_status[k]:>6,}")
    print(f"\n  distinct (code, sheet, range) with MISMATCH: {len(by_code_range)}")

    # ---- render report ----
    L = [
        "# Moldova formula parsing verification\n",
        "Cross-checks every TAG rule in `data/tag_rules.csv` by applying its "
        "parsed `criteria_json` to the matching year-range raw sheet and "
        "comparing the per-year sum against Excel's cached cell value in the "
        "Approved / Executed sheet. A MATCH confirms that our parsed criteria "
        "reproduce Excel exactly; a MISMATCH points at a parser gap, a "
        "formula feature we don't yet model (IF-wrapping, cell arithmetic, "
        "INDIRECT), or a hand-entered override.\n",
        "\n## Summary\n",
        "| Status | Count |",
        "|---|---:|",
    ]
    for k in ("MATCH", "MISMATCH", "BOTH_EMPTY", "EXCEL_EMPTY",
             "EXCEL_EMPTY_PIPELINE_NONZERO", "PIPELINE_NONE",
             "SKIP_UNSUPPORTED_MEASURE"):
        if by_status.get(k):
            L.append(f"| `{k}` | {by_status[k]:,} |")
    L.append("")
    L.append(f"Tolerance: |pipeline − excel| ≤ {MATCH_EPSILON} currency units.\n")

    if by_code_range:
        L.append("## Rules with at least one mismatch\n")
        L.append("One section per (code, sheet, range). Each shows the formula the "
                 "extractor captured and the per-year diff for years where pipeline "
                 "and excel disagree.\n")
        for (code, sheet, rk), rows in sorted(by_code_range.items(),
                                              key=lambda kv: (-len(kv[1]), kv[0])):
            rows.sort(key=lambda d: d["year"] or 0)
            L.append(f"### `{code}` — {sheet}, range={rk} "
                     f"({len(rows)} mismatched year(s))\n")
            L.append(f"Captured formula (sample year): `{rows[0]['sample_formula']}`\n")
            L.append(f"`n_sumifs = {rows[0]['n_sumifs']}`\n")
            L.append("| year | excel | pipeline | diff | rel |")
            L.append("|---:|---:|---:|---:|---:|")
            for d in rows:
                excel = d["excel"] or 0
                pipe = d["pipeline"] or 0
                diff = d["diff"] or 0
                rel = f"{100 * diff / excel:.2f}%" if excel else "—"
                L.append(f"| {d['year']} | {excel:,.0f} | {pipe:,.0f} | "
                         f"{diff:,.0f} | {rel} |")
            L.append("")

    L.append("## Known unmodeled formula features\n")
    L.append("Two categories of mismatch are expected and documented here so "
             "the SME can decide whether to accept or resolve upstream:\n")
    L.append("- **Cell-subtraction** — `=SUMIFS(…) - C19`. The SUMIFS result "
             "is adjusted by subtracting another cell's value (usually a "
             "sibling tag-row's override). The parser captures the SUMIFS "
             "but discards the trailing `-cell_ref`, so pipeline > excel by "
             "that cell's value. Affects `SOC_ASS`, `PUB_SAF`, `SOC_PRO`.\n")
    L.append("- **Raw2 supplements** — `=SUMIFS(…) + SUM(SUMIFS('Raw2'!$F:$F, …))`. "
             "The second SUMIFS pulls from the hidden `Raw2` sheet, which "
             "uses a different schema (`admin6` instead of `admin1`, no "
             "`econ0`). Our silver mapping doesn't include Raw2, so that "
             "branch's contribution is dropped. Affects `WAT_SAN`.\n")
    (REPORTS / "parsing_verification.md").write_text("\n".join(L))
    print(f"\nWrote reports/parsing_verification.md "
          f"({len(mismatches)} mismatched (rule, year) cells across "
          f"{len(by_code_range)} rule(s))")


if __name__ == "__main__":
    verify()
