"""Generate ISSUES.md — known formula-level problems in the Moldova workbook.

Reports (does not fix) anything that would force the pipeline to deviate
from a literal translation of the Excel formulas:

  1. Hard-coded overrides — literal values in cells where the column is otherwise SUMIFS.
  2. Broken named ranges (#REF! targets) and any formula that uses them.
  3. External references — formulas pointing to files outside this workbook.
  4. Inconsistent formulas within the same column of Approved/Executed.
  5. Approved vs Executed structural diffs (rows/cols where one has a formula and the other does not).
  6. Volatile / non-deterministic functions (INDIRECT, OFFSET, TODAY, NOW, RAND).

Inputs: temp/Moldova BOOST.xlsx, defined_names.csv, formula_map.csv, hardcoded_columns.csv, hardcoded_cells.csv
Output: Moldova/_analysis/ISSUES.md
"""
from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent          # Moldova/_onboarding
COUNTRY_DIR = ROOT.parent                              # Moldova/
XLSX = COUNTRY_DIR.parent / "temp" / "Moldova BOOST.xlsx"
DATA = ROOT / "data"
REPORTS = ROOT / "reports"
REPORTS.mkdir(parents=True, exist_ok=True)
OUT = DATA  # load_csv() reads inputs from data/ (tag_rules.csv lives at COUNTRY_DIR)

VOLATILE = re.compile(r"\b(INDIRECT|OFFSET|TODAY|NOW|RAND|RANDBETWEEN)\s*\(", re.I)
EXTERNAL_REF = re.compile(r"\[[^\]]+\.xls[xm]?\]")  # [OtherFile.xlsx]Sheet!A1
REF_ERR = re.compile(r"#REF!")


def normalize_formula_shape(formula: str) -> str:
    """Replace cell coordinates with placeholders so two structurally identical
    formulas in different rows compare equal. e.g.,
      =SUMIFS(approved,Exclude,"<>y",year,O$1)  ->  =SUMIFS(approved,Exclude,"<>y",year,#$1)
    """
    # column-letter + row -> #
    s = re.sub(r"\$?[A-Z]{1,3}\$?\d+", "#", formula)
    s = re.sub(r"\s+", "", s)
    return s


# Measure-agnostic normalizer: collapse approved/executed/revised (including
# country-specific suffixed variants like `approved_16`, `executed_20`, or the
# `_r` revenue counterparts) to a common token so an Approved-sheet SUMIFS
# compares equal to the matching Executed-sheet SUMIFS when only the measure
# name differs. The optional `_\w+` tail covers suffixes like `_16`, `_20`, `_r`.
_MEASURE_RE = re.compile(r"\b(approved_orig|approved|executed|revised)(_\w+)?\b")


def normalize_measure_agnostic(formula: str) -> str:
    s = normalize_formula_shape(formula)
    return _MEASURE_RE.sub("MEASURE", s)


def load_csv(name: str) -> list[dict]:
    """Look first at the country root (DLT drivers: tag_rules.csv,
    code_dictionary.csv) then the onboarding data dir (everything else)."""
    for base in (COUNTRY_DIR, OUT):
        p = base / name
        if p.exists():
            with p.open() as f:
                return list(csv.DictReader(f))
    return []


def main():
    defined_names = {r["name"]: r["target"] for r in load_csv("defined_names.csv")}
    formulas = load_csv("formula_map.csv")
    hardcoded_cols = load_csv("hardcoded_columns.csv")
    hardcoded_cells = load_csv("hardcoded_cells.csv")

    issues: dict[str, list[str]] = defaultdict(list)

    # 1. Hard-coded overrides — summarise from hardcoded_columns.csv (severity by count)
    # Build the set of (sheet, col_idx) considered NO_DATA so we can skip them everywhere.
    no_data_cols = {(r["sheet"], int(r["col_idx"])) for r in hardcoded_cols if r["classification"] == "NO_DATA"}
    # Header lookup: prefer human-readable column header (mostly years) over col letter.
    col_header = {(r["sheet"], int(r["col_idx"])): (r["header"] or r["col_letter"]) for r in hardcoded_cols}

    col_overrides = []
    for r in hardcoded_cols:
        if r["classification"] == "MIXED_LITERAL_FORMULA" and int(r["n_literal_numeric"]) > 0:
            if int(r["n_formula"]) == 0:
                continue
            col_overrides.append(r)
    col_overrides.sort(key=lambda r: int(r["n_literal_numeric"]), reverse=True)

    # 2. Broken named ranges
    broken_names = {n: t for n, t in defined_names.items() if REF_ERR.search(t or "")}
    formulas_using_broken = []
    if broken_names:
        for f in formulas:
            ranges = (f["named_ranges"] or "").split("|") if f["named_ranges"] else []
            hit = [r for r in ranges if r in broken_names]
            if hit:
                formulas_using_broken.append((f, hit))

    # 3. External references — scan raw formula strings (need workbook pass since formula_map already has them)
    external_refs = [f for f in formulas if EXTERNAL_REF.search(f["formula"])]

    # 4. Inconsistent formulas within the same column of Approved/Executed
    by_sheet_col: dict[tuple[str, int], list[dict]] = defaultdict(list)
    for f in formulas:
        if int(f["row"]) == 1:  # skip header row
            continue
        by_sheet_col[(f["sheet"], int(f["col"]))].append(f)

    inconsistent_cols = []
    for (sheet, col_idx), cells in by_sheet_col.items():
        if (sheet, col_idx) in no_data_cols:
            continue  # year column with no underlying data — not an issue
        shapes = Counter(normalize_formula_shape(c["formula"]) for c in cells)
        if len(shapes) > 1:
            # ignore if dominant shape > 95% (single outlier is informative but not "inconsistent column")
            top, top_n = shapes.most_common(1)[0]
            total = sum(shapes.values())
            inconsistent_cols.append({
                "sheet": sheet,
                "col_letter": get_column_letter(col_idx),
                "col_idx": col_idx,
                "n_shapes": len(shapes),
                "n_cells": total,
                "top_shape_pct": round(100 * top_n / total, 1),
                "shapes": shapes,
            })
    inconsistent_cols.sort(key=lambda r: (-r["n_shapes"], r["sheet"], r["col_idx"]))

    # 5. Approved vs Executed structural diff — cells where one has formula and other does not (same coord)
    approved = {(int(f["row"]), int(f["col"])): f for f in formulas if f["sheet"] == "Approved"}
    executed = {(int(f["row"]), int(f["col"])): f for f in formulas if f["sheet"] == "Executed"}
    only_in_approved = sorted(set(approved) - set(executed))
    only_in_executed = sorted(set(executed) - set(approved))
    # also: same coord, different shape
    shape_diffs = []
    measure_only_diffs = 0
    for k in set(approved) & set(executed):
        af = approved[k]["formula"]
        ef = executed[k]["formula"]
        sa = normalize_formula_shape(af)
        se = normalize_formula_shape(ef)
        if sa == se:
            continue
        # Ignore differences that are only the measure name (approved vs executed).
        if normalize_measure_agnostic(af) == normalize_measure_agnostic(ef):
            measure_only_diffs += 1
            continue
        shape_diffs.append((k, af, ef))

    # 6. Volatile functions
    volatile = [f for f in formulas if VOLATILE.search(f["formula"])]

    # ---------- write ISSUES.md ----------
    L: list[str] = []
    L.append("# Moldova BOOST — Known Formula Issues\n")
    L.append("Generated by `phase1_issues.py`. **Report-only — the pipeline translates formulas literally; nothing here is fixed in code.** Each item is for the subject-matter expert to triage.\n")

    L.append(f"_Skipped {len(no_data_cols)} column(s) classified `NO_DATA` (formula present but no underlying data — typically years not yet reported)._\n")
    L.append("## 1. Hard-coded overrides (literal values in formula columns)\n")
    override_cells = [
        c for c in hardcoded_cells
        if (c["sheet"], int(c["col_idx"])) not in no_data_cols
    ]
    if override_cells:
        override_cells.sort(key=lambda c: (c["sheet"], c["header"], int(c["row"])))
        L.append(
            f"**{len(override_cells)} cells** in Approved/Executed contain hand-entered "
            "numeric values instead of a SUMIFS formula. The pipeline reproduces SUMIFS "
            "results only; each override below will surface as a discrepancy.\n"
        )
        L.append("| Sheet | Year | Cell | Code | Category | Value |")
        L.append("|---|---|---|---|---|---:|")
        for c in override_cells[:200]:
            try:
                val = f"{float(c['value']):,.0f}"
            except (TypeError, ValueError):
                val = c["value"]
            L.append(
                f"| {c['sheet']} | {c['header']} | {c['cell']} | "
                f"`{c.get('code') or '—'}` | {c.get('category') or '—'} | {val} |"
            )
        if len(override_cells) > 200:
            L.append(f"\n_…and {len(override_cells)-200} more. Full list: `hardcoded_cells.csv`._")
    else:
        L.append("_None._")

    L.append("\n## 2. Broken named ranges (`#REF!`)\n")
    if broken_names:
        for n, t in broken_names.items():
            L.append(f"- `{n}` → `{t}`")
        L.append(f"\n**Formulas referencing broken names: {len(formulas_using_broken)}**")
        for f, hit in formulas_using_broken[:20]:
            L.append(f"  - {f['sheet']}!{f['cell']} uses {hit}: `{f['formula'][:120]}`")
        if len(formulas_using_broken) > 20:
            L.append(f"  - …and {len(formulas_using_broken)-20} more.")
    else:
        L.append("_None._")

    L.append("\n## 3. External-workbook references\n")
    if external_refs:
        for f in external_refs[:50]:
            L.append(f"- {f['sheet']}!{f['cell']}: `{f['formula'][:160]}`")
        if len(external_refs) > 50:
            L.append(f"\n_…and {len(external_refs)-50} more._")
    else:
        L.append("_None._")

    L.append("\n## 4. Approved vs Executed structural differences\n")
    L.append(f"- Cells with formula in **Approved only**: {len(only_in_approved)}")
    L.append(f"- Cells with formula in **Executed only**: {len(only_in_executed)}")
    L.append(f"- Cells with formulas in both but **different shape** (excluding measure-only diffs): {len(shape_diffs)}")
    L.append(f"- _Ignored: {measure_only_diffs} cells where the only difference was the measure name (e.g. `approved` vs `executed`, `approved_20` vs `executed_20`) — that's expected._")
    if shape_diffs:
        # Attach the code for each diff so the SME can see which tag is affected.
        code_by_row: dict[int, str] = {}
        for f in formulas:
            if int(f.get("col", 0)) == 1 and f.get("formula"):  # column A usually holds codes, but codes aren't formulas
                continue
        # Pull codes from tag_rules.csv (stable lookup sheet × row → code).
        tag_rules = load_csv("tag_rules.csv")
        code_lookup: dict[tuple[str, int], str] = {}
        for r in tag_rules:
            try:
                code_lookup[(r["sheet"], int(r["row"]))] = r["code"]
            except (KeyError, ValueError):
                pass

        L.append("\nFirst 20 shape differences (full formulas):\n")
        for (r, c), af, ef in shape_diffs[:20]:
            label = col_header.get(("Approved", c)) or col_header.get(("Executed", c)) or get_column_letter(c)
            code = code_lookup.get(("Approved", r)) or code_lookup.get(("Executed", r)) or ""
            code_str = f" — `{code}`" if code else ""
            L.append(f"- **Row {r}, {label}**{code_str}")
            L.append("  ```")
            L.append(f"  Approved: {af}")
            L.append(f"  Executed: {ef}")
            L.append("  ```")
        if len(shape_diffs) > 20:
            L.append(f"\n_…and {len(shape_diffs)-20} more._")

    L.append("\n## 5. Volatile / non-deterministic functions\n")
    if volatile:
        L.append(f"**{len(volatile)} formulas** use INDIRECT/OFFSET/TODAY/NOW/RAND. These cannot be translated to deterministic Spark and need expert resolution.\n")
        for f in volatile[:30]:
            L.append(f"- {f['sheet']}!{f['cell']}: `{f['formula'][:160]}`")
        if len(volatile) > 30:
            L.append(f"\n_…and {len(volatile)-30} more._")
    else:
        L.append("_None._")

    L.append("\n---")
    L.append("\n## Triage workflow\n")
    L.append("1. Walk each section with the subject-matter expert.")
    L.append("2. For each item: decide **fix-Excel** (correct the workbook upstream) or **accept-as-is** (pipeline reports the divergence in the per-run discrepancy review).")
    L.append("3. Re-run `phase1_issues.py` after the workbook is updated; resolved items disappear automatically.")

    provenance = [
        "> Generated by `scripts/4_generate_issues.py` from",
        "> `data/defined_names.csv`, `data/formula_map.csv`, `data/hardcoded_columns.csv`, `data/hardcoded_cells.csv`.",
        "> Section 6 (suspicious tag rules) is appended by `scripts/5_build_code_dictionary.py`.",
        "> Rerun with `python3 scripts/4_generate_issues.py && python3 scripts/5_build_code_dictionary.py`.\n",
    ]
    (REPORTS / "ISSUES.md").write_text("\n".join(provenance + L))
    print(f"Wrote reports/ISSUES.md")
    print(f"  hardcoded override columns: {len(col_overrides)}")
    print(f"  broken named ranges: {len(broken_names)} (formulas using them: {len(formulas_using_broken)})")
    print(f"  external refs: {len(external_refs)}")
    print(f"  inconsistent columns: {len(inconsistent_cols)}")
    print(f"  approved-only cells: {len(only_in_approved)}, executed-only: {len(only_in_executed)}, shape diffs: {len(shape_diffs)} (ignored measure-only: {measure_only_diffs})")
    print(f"  volatile formulas: {len(volatile)}")


if __name__ == "__main__":
    main()
