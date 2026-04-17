"""Extract a tag-rules table from Approved/Executed.

Each row in those sheets is a BOOST classification tag. Its SUMIFS criteria
define the filter on the Expenditure raw data. We parse those criteria
verbatim — no interpretation — into a structured table that the silver layer
will iterate over.

Output:
  tag_rules.csv  one row per Approved-sheet (or Executed-sheet) row, with:
    sheet, row, code, category, fiscal_year, region, income, basis, source,
    rigidity, formulas_text, coverage_text, sample_formula, sample_year,
    measure, criteria_json

The `criteria_json` is a JSON list of {field, op, value} dicts, e.g.:
  [{"field":"Exclude","op":"<>","value":"y"},
   {"field":"econ3","op":"=","value":"606"}]

Some countries hand-code their classification logic in PySpark; where the
workbook itself encodes the classification as SUMIFS criteria (the Zimbabwe
pattern reused here), this script lifts that encoding so the pipeline can
apply it literally.
"""
from __future__ import annotations

import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

XLSX = Path("/Users/ysuzuki2/Developer/mega-boost/temp/Moldova BOOST.xlsx")
ROOT = Path("/Users/ysuzuki2/Developer/mega-boost/Moldova/_analysis")
OUT = ROOT / "data"
OUT.mkdir(parents=True, exist_ok=True)

TARGET_SHEETS = ("Approved", "Executed")

# Column map for Approved/Executed rows (1-indexed cells).
COLS = {
    "code":         "A",
    "category":     "B",
    "formulas":     "AE",
    "coverage":     "AF",
    "fiscal_year":  "AG",
    "region":       "AH",
    "income":       "AI",
    "basis":        "AJ",
    "source":       "AK",
    "rigidity":     "AL",
}

# Year header columns: scan row 1 cached values; treat 4-digit numbers as years.
YEAR_HEADER_RE = re.compile(r"^\s*(19|20)\d{2}\s*$")


def split_sumifs_args(inner: str) -> list[str]:
    """Split the inside of a SUMIFS(...) call by top-level commas, respecting
    quoted strings and nested parens/braces. Array constants like
    {"a","b","c"} must not split on their internal commas."""
    out, buf, depth, in_str = [], [], 0, False
    for ch in inner:
        if ch == '"' and (not buf or buf[-1] != "\\"):
            in_str = not in_str
            buf.append(ch)
            continue
        if in_str:
            buf.append(ch)
            continue
        if ch in "({":
            depth += 1
            buf.append(ch)
        elif ch in ")}":
            depth -= 1
            buf.append(ch)
        elif ch == "," and depth == 0:
            out.append("".join(buf).strip())
            buf = []
        else:
            buf.append(ch)
    if buf:
        out.append("".join(buf).strip())
    return out


def parse_criterion(value: str) -> tuple[str, str]:
    """Return (op, literal). Excel criteria can be:
      "<>y", ">=10", "606", "<2024", year-cell-ref like O$1.
    Cell refs are returned as op="=year" so the caller can substitute the year.
    """
    v = value.strip()
    # quoted string
    if v.startswith('"') and v.endswith('"'):
        v = v[1:-1]
        for op in ("<>", "<=", ">=", "<", ">", "="):
            if v.startswith(op):
                return op, v[len(op):]
        return "=", v
    # cell reference (year header lookup) — column letter + $row+digits
    if re.match(r"^\$?[A-Z]{1,3}\$?\d+$", v):
        return "=year", v
    # numeric / unquoted literal
    return "=", v


# Match SUMIFS(...) including nested parens.
SUMIFS_RE = re.compile(r"SUMIFS\s*\(", re.I)


def extract_sumifs_blocks(formula: str) -> list[str]:
    """Return every top-level SUMIFS(...) inner-arg-string in the formula."""
    blocks = []
    i = 0
    while True:
        m = SUMIFS_RE.search(formula, i)
        if not m:
            break
        depth = 1
        j = m.end()
        while j < len(formula) and depth:
            if formula[j] == "(":
                depth += 1
            elif formula[j] == ")":
                depth -= 1
            j += 1
        blocks.append(formula[m.end(): j - 1])
        i = j
    return blocks


def _split_array_constant(raw: str) -> list[str]:
    """Split `{"a","b","c"}` into its string elements, respecting quoted commas."""
    inner = raw.strip()[1:-1]  # drop { }
    parts = []
    buf = []
    in_str = False
    for ch in inner:
        if ch == '"':
            in_str = not in_str
            buf.append(ch)
        elif ch == "," and not in_str:
            parts.append("".join(buf).strip())
            buf = []
        else:
            buf.append(ch)
    if buf:
        parts.append("".join(buf).strip())
    return parts


def _parse_one_block(block: str) -> tuple[str, list[list[dict]]]:
    """Parse a single SUMIFS(...) block. Each (field, value-or-array) pair
    becomes one or more criteria options; array constants like
    `{"A","B","C"}` expand to N alternatives. The cartesian product across
    criteria yields a list of criterion-lists — `SUMIFS(…, econ2, {"A","B"})`
    becomes two branches (one with `econ2 = "A"`, one with `econ2 = "B"`)
    because Excel applies SUMIFS once per array element and the caller sums
    the results.

    Returns (measure, branches). An empty `branches` list means the block
    couldn't be parsed cleanly."""
    from itertools import product

    args = split_sumifs_args(block)
    if len(args) < 3:
        return args[0] if args else "", []
    measure = args[0]

    per_criterion: list[list[dict]] = []
    for k in range(1, len(args) - 1, 2):
        field = args[k].strip()
        raw_value = args[k + 1].strip()
        if raw_value.startswith("{") and raw_value.endswith("}"):
            elements = _split_array_constant(raw_value)
            opts = []
            for elem in elements:
                op, lit = parse_criterion(elem)
                opts.append({"field": field, "op": op, "value": lit})
            per_criterion.append(opts)
        else:
            op, lit = parse_criterion(raw_value)
            per_criterion.append([{"field": field, "op": op, "value": lit}])

    branches = [list(combo) for combo in product(*per_criterion)]
    return measure, branches


def parse_criteria(formula: str) -> tuple[str, list[list[dict]], list[str]]:
    """See the original docstring for the OR-across-blocks semantics.

    Returns `(primary_measure, criteria_groups, measures_per_block)` where
    `criteria_groups` is a FLAT list of all branches (block-level OR + any
    array-constant expansion). A raw row satisfies the rule when any single
    inner list's criteria all match."""
    blocks = extract_sumifs_blocks(formula)
    if not blocks:
        return "", [], []
    groups: list[list[dict]] = []
    measures: list[str] = []
    for blk in blocks:
        m, branches = _parse_one_block(blk)
        if branches:
            groups.extend(branches)
        measures.append(m)
    primary = measures[0] if measures else ""
    return primary, groups, measures


# Shape-normalizer: replaces year-header cell refs (like M$1, N$1) with YEAR so
# formulas that differ only by which year column they pull from collapse to one
# shape. Other cell refs are preserved — the structure of the rule (criteria
# count, field names, filter values) is part of the shape key.
_YEAR_CELL_RE = re.compile(r"\$?[A-Z]{1,3}\$?1\b")


def normalize_formula_shape(formula: str) -> str:
    s = _YEAR_CELL_RE.sub("YEAR", formula)
    return re.sub(r"\s+", "", s)


def main():
    wb_f = load_workbook(XLSX, read_only=True, data_only=False)
    wb_v = load_workbook(XLSX, read_only=True, data_only=True)

    out_rows = []
    for sheet in TARGET_SHEETS:
        if sheet not in wb_f.sheetnames:
            continue
        ws_f = wb_f[sheet]
        ws_v = wb_v[sheet]
        v_grid = list(ws_v.iter_rows(values_only=True))
        f_grid = list(ws_f.iter_rows(values_only=False))

        # Identify year-header columns from row 1 cached values.
        year_cols: list[tuple[int, int]] = []  # (col_idx, year)
        if v_grid:
            for c_idx, hv in enumerate(v_grid[0], start=1):
                if hv is None:
                    continue
                hv_str = str(int(hv)) if isinstance(hv, float) and hv.is_integer() else str(hv).strip()
                if YEAR_HEADER_RE.match(hv_str):
                    year_cols.append((c_idx, int(hv_str)))

        col_idx = {k: column_index_from_string(v) for k, v in COLS.items()}

        def vget(r: int, c: int):
            if r - 1 >= len(v_grid):
                return None
            row_v = v_grid[r - 1]
            return row_v[c - 1] if c - 1 < len(row_v) else None

        for r_idx, row in enumerate(f_grid, start=1):
            if r_idx == 1:
                continue
            code_cell = vget(r_idx, col_idx["code"])
            if code_cell is None or str(code_cell).strip() == "":
                continue

            # Group year columns by their normalized formula shape. Each distinct
            # shape becomes one rule row — so a tag code with different formulas
            # for 2006-15 vs 2016-19 vs 2020-24 emits three rules, each covering
            # its own year span. This handles Moldova's per-year-range schema
            # switches (different named ranges, different criteria fields/values,
            # different filter languages, extra Raw2 additions, etc.).
            # Index into the already-materialized f_grid rather than calling
            # ws_f.cell(), which is O(sheet) per call in openpyxl read-only mode.
            row_cells = f_grid[r_idx - 1] if r_idx - 1 < len(f_grid) else ()
            shape_groups: dict[str, dict] = {}
            for c_idx, year in year_cols:
                if c_idx - 1 >= len(row_cells):
                    continue
                cell = row_cells[c_idx - 1].value
                if not isinstance(cell, str) or not cell.startswith("="):
                    continue
                shape = normalize_formula_shape(cell)
                g = shape_groups.setdefault(shape, {
                    "years": [], "sample_formula": cell, "sample_year": year,
                    "has_sumifs": "SUMIFS" in cell.upper(),
                })
                g["years"].append(year)

            meta_cells = {
                "category":      vget(r_idx, col_idx["category"]),
                "fiscal_year":   vget(r_idx, col_idx["fiscal_year"]),
                "region":        vget(r_idx, col_idx["region"]),
                "income":        vget(r_idx, col_idx["income"]),
                "basis":         vget(r_idx, col_idx["basis"]),
                "source":        vget(r_idx, col_idx["source"]),
                "rigidity":      vget(r_idx, col_idx["rigidity"]),
                "formulas_text": vget(r_idx, col_idx["formulas"]),
                "coverage_text": vget(r_idx, col_idx["coverage"]),
            }

            if not shape_groups:
                # Row has a code but no formula in any year column — HEADER.
                out_rows.append({
                    "sheet": sheet, "row": r_idx, "row_type": "HEADER",
                    "code": code_cell, **meta_cells,
                    "years_covered": "", "year_min": None, "year_max": None,
                    "sample_formula": "", "sample_year": None,
                    "measure": "", "criteria_json": "[]",
                    "block_measures": "", "n_sumifs": 0,
                })
                continue

            # Emit one row per distinct shape group.
            for shape, g in sorted(shape_groups.items(), key=lambda kv: min(kv[1]["years"])):
                row_type = "TAG" if g["has_sumifs"] else "ROLLUP"
                if row_type == "TAG":
                    measure, criteria_groups, block_measures = parse_criteria(g["sample_formula"])
                else:
                    measure, criteria_groups, block_measures = "", [], []
                years_sorted = sorted(g["years"])
                # criteria_json is a list-of-lists: each inner list is one
                # SUMIFS block's criteria. A raw row matches the rule if ANY
                # inner list's criteria all match (SUMIFS+SUMIFS = OR).
                n_sumifs = len(block_measures)
                out_rows.append({
                    "sheet":          sheet,
                    "row":            r_idx,
                    "row_type":       row_type,
                    "code":           code_cell,
                    **meta_cells,
                    "years_covered":  ",".join(str(y) for y in years_sorted),
                    "year_min":       years_sorted[0],
                    "year_max":       years_sorted[-1],
                    "sample_formula": g["sample_formula"],
                    "sample_year":    g["sample_year"],
                    "measure":        measure,
                    "criteria_json":  json.dumps(criteria_groups, ensure_ascii=False),
                    "block_measures": "|".join(block_measures),
                    "n_sumifs":       n_sumifs,
                })

    headers = list(out_rows[0].keys()) if out_rows else []
    with (OUT / "tag_rules.csv").open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        w.writerows(out_rows)

    # Console summary
    print(f"Wrote {len(out_rows)} tag rules → tag_rules.csv")
    by_sheet_type = {}
    by_field_count = {}
    fields_used = {}
    measures = {}
    for r in out_rows:
        key = (r["sheet"], r["row_type"])
        by_sheet_type[key] = by_sheet_type.get(key, 0) + 1
        if r["row_type"] != "TAG":
            continue
        groups = json.loads(r["criteria_json"])
        # Distribution key: criterion count of the first (primary) SUMIFS block.
        primary = groups[0] if groups else []
        by_field_count[len(primary)] = by_field_count.get(len(primary), 0) + 1
        for c in primary:
            fields_used[c["field"]] = fields_used.get(c["field"], 0) + 1
        measures[r["measure"]] = measures.get(r["measure"], 0) + 1
    print("  rows by sheet × type:")
    for k, v in sorted(by_sheet_type.items()):
        print(f"    {k[0]:<10} {k[1]:<8} {v}")
    print(f"  TAG rows criterion-count distribution: {dict(sorted(by_field_count.items()))}")
    print(f"  TAG fields used in criteria: {sorted(fields_used.items(), key=lambda x: -x[1])}")
    print(f"  TAG measures summed: {measures}")


if __name__ == "__main__":
    main()
