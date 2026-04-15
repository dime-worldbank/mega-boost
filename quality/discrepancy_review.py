"""Country-agnostic utility for generating a per-country discrepancy review doc.

The previous bottleneck for every country was the back-and-forth with the
subject-matter expert about why pipeline aggregates differ from the Excel
"Approved"/"Executed" totals (Excel often double-counts; the pipeline is
strict). This module turns that loop into a single regenerable markdown doc:

    quality/_reviews/{country_code}_discrepancy_review.md

Expert resolutions written into the doc survive across regenerations (keyed by
dimension hash), so the doc accumulates institutional knowledge instead of
being thrown away each iteration.

Public API
----------
    load_excel_totals(xlsx_path, sheet, dimension_cols, value_cols)
        -> pandas.DataFrame  (the Excel side, with cached values)

    load_formula_map(xlsx_path, target_sheets)
        -> pandas.DataFrame  (formula text per cell — used for traceability)

    build_review(
        country_code,
        excel_path,
        pipeline_df,           # pandas or pyspark DataFrame
        dimension_cols,        # e.g. ["year", "econ1"]
        value_cols,            # e.g. ["approved", "executed"]
        excel_sheet="Executed",
        threshold=0.05,        # 5% acceptable discrepancy
        formula_map=None,      # optional pre-loaded formula map
        out_path=None,         # default: quality/_reviews/{code}_discrepancy_review.md
        pipeline_source_ref="", # e.g. "Zimbabwe/ZWE_transform_load_raw_dlt.py:silver"
    ) -> Path

    classify_discrepancy(excel_value, pipeline_value, formula_text="") -> str
        Heuristic tag: double-count / triple-count / overlap / missing-filter /
        rounding / unknown.

    merge_expert_resolutions(old_md, new_md) -> str
        Preserves prior `### Expert resolution` blocks keyed by dimension hash.
"""
from __future__ import annotations

import csv
import hashlib
import re
from pathlib import Path
from typing import Iterable

# Pandas is always available in this project; openpyxl is a notebook-time dep.
import pandas as pd

REVIEWS_DIR = Path(__file__).resolve().parent / "_reviews"
RESOLUTION_HEADING = "### Expert resolution"
DIMENSION_KEY_PREFIX = "<!-- dim-key: "
DIMENSION_KEY_SUFFIX = " -->"


# ---------- dimension hashing ----------

def _dim_key(dim_values: Iterable) -> str:
    s = "|".join("" if v is None else str(v) for v in dim_values)
    return hashlib.sha1(s.encode()).hexdigest()[:12]


# ---------- excel side ----------

def load_excel_totals(xlsx_path, sheet: str, dimension_cols: list[str],
                      value_cols: list[str]) -> pd.DataFrame:
    """Read the Excel sheet's cached values and aggregate by dimension cols.

    `..` (the BOOST missing-value marker) is treated as NaN.
    Years that are dimension keys but stored as column headers (wide format)
    must be unpivoted by the caller before passing — this helper assumes the
    Excel sheet is already long-form keyed by `dimension_cols`.
    """
    df = pd.read_excel(xlsx_path, sheet_name=sheet, na_values=["..", ".", "n/a", "N/A"])
    keep = [c for c in dimension_cols + value_cols if c in df.columns]
    missing = set(dimension_cols + value_cols) - set(keep)
    if missing:
        raise KeyError(f"Excel sheet {sheet!r} missing columns: {sorted(missing)}")
    out = df[keep].dropna(how="all", subset=value_cols)
    out = out.groupby(dimension_cols, dropna=False, as_index=False)[value_cols].sum(min_count=1)
    return out


def load_formula_map(xlsx_path, target_sheets=("Approved", "Executed")) -> pd.DataFrame:
    """Per-cell formula extractor. Returns columns:
       sheet, cell, row, col, formula, cached_value.
    Lifted from Zimbabwe/_analysis/phase1_extract.py for reuse.
    """
    from openpyxl import load_workbook  # notebook-time import
    wb_f = load_workbook(xlsx_path, read_only=True, data_only=False)
    wb_v = load_workbook(xlsx_path, read_only=True, data_only=True)
    rows = []
    for sheet in target_sheets:
        if sheet not in wb_f.sheetnames:
            continue
        ws_f = wb_f[sheet]
        v_grid = list(wb_v[sheet].iter_rows(values_only=True))
        for r_idx, row in enumerate(ws_f.iter_rows(values_only=False), start=1):
            for c_idx, cell in enumerate(row, start=1):
                v = cell.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                cached = None
                if r_idx - 1 < len(v_grid) and c_idx - 1 < len(v_grid[r_idx - 1]):
                    cached = v_grid[r_idx - 1][c_idx - 1]
                rows.append({"sheet": sheet, "cell": cell.coordinate,
                             "row": r_idx, "col": c_idx,
                             "formula": v, "cached_value": cached})
    return pd.DataFrame(rows)


# ---------- diff & classify ----------

_DOUBLE_COUNT_HINTS = re.compile(
    r"(capital.*social|social.*capital|total.*sub|sub.*total|all.*excl)",
    re.I,
)


def classify_discrepancy(excel_value, pipeline_value, formula_text: str = "") -> str:
    if excel_value is None or pd.isna(excel_value):
        return "excel-missing"
    if pipeline_value is None or pd.isna(pipeline_value):
        return "pipeline-missing"
    e, p = float(excel_value), float(pipeline_value)
    if p == 0:
        return "pipeline-zero"
    ratio = e / p
    # Common Excel double/triple-count signature: excel = N * pipeline
    if 1.95 <= ratio <= 2.05:
        return "double-count"
    if 2.95 <= ratio <= 3.05:
        return "triple-count"
    if 0.99 <= ratio <= 1.01:
        return "rounding"
    if formula_text and _DOUBLE_COUNT_HINTS.search(formula_text):
        return "overlap"
    if abs(e - p) / max(abs(e), 1) < 0.05:
        return "minor-drift"
    return "unknown"


def _pct(e, p) -> float:
    if e is None or pd.isna(e) or p is None or pd.isna(p):
        return float("nan")
    base = max(abs(float(e)), abs(float(p)), 1.0)
    return abs(float(e) - float(p)) / base


def _diff(excel_df: pd.DataFrame, pipeline_df: pd.DataFrame,
          dimension_cols: list[str], value_cols: list[str], threshold: float):
    merged = excel_df.merge(
        pipeline_df, on=dimension_cols, how="outer",
        suffixes=("_excel", "_pipeline"),
    )
    rows = []
    for _, row in merged.iterrows():
        for vc in value_cols:
            e = row.get(f"{vc}_excel")
            p = row.get(f"{vc}_pipeline")
            pct = _pct(e, p)
            if pd.isna(pct) or pct > threshold:
                rows.append({
                    "dim_key": _dim_key([row[c] for c in dimension_cols] + [vc]),
                    **{c: row[c] for c in dimension_cols},
                    "measure": vc,
                    "excel": e,
                    "pipeline": p,
                    "delta_abs": (None if pd.isna(pct) else float(e) - float(p)),
                    "delta_pct": (None if pd.isna(pct) else pct),
                })
    return pd.DataFrame(rows)


# ---------- markdown rendering ----------

def _fmt_num(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "—"
    if isinstance(v, (int, float)):
        return f"{v:,.2f}" if abs(v) < 1e15 else f"{v:.4e}"
    return str(v)


def _render(country_code: str, excel_path, dimension_cols: list[str],
            value_cols: list[str], threshold: float,
            diff_df: pd.DataFrame, formula_map: pd.DataFrame | None,
            pipeline_source_ref: str) -> str:
    L: list[str] = []
    L.append(f"# {country_code.upper()} — Pipeline vs. Excel Discrepancy Review\n")
    L.append(f"Source workbook: `{Path(excel_path).name}`")
    L.append(f"Pipeline source: `{pipeline_source_ref or '(unspecified)'}`")
    L.append(f"Threshold: {threshold:.0%}\n")

    if diff_df.empty:
        L.append("**No discrepancies above threshold.** Pipeline matches Excel within tolerance.\n")
        return "\n".join(L)

    L.append(f"## Summary\n")
    L.append(f"- Total flagged rows: **{len(diff_df)}**")
    cause_counts = diff_df.assign(
        cause=diff_df.apply(lambda r: classify_discrepancy(r["excel"], r["pipeline"]), axis=1)
    )["cause"].value_counts()
    for cause, n in cause_counts.items():
        L.append(f"  - `{cause}`: {n}")

    L.append("\n## Discrepancies\n")
    for _, r in diff_df.iterrows():
        cause = classify_discrepancy(r["excel"], r["pipeline"])
        dim_str = " · ".join(f"**{c}**=`{r[c]}`" for c in dimension_cols)
        L.append(f"### {dim_str} · **{r['measure']}**  _[{cause}]_")
        L.append(f"{DIMENSION_KEY_PREFIX}{r['dim_key']}{DIMENSION_KEY_SUFFIX}")
        L.append("")
        L.append(f"| Side | Value |")
        L.append(f"|---|---:|")
        L.append(f"| Excel    | {_fmt_num(r['excel'])} |")
        L.append(f"| Pipeline | {_fmt_num(r['pipeline'])} |")
        L.append(f"| Δ abs    | {_fmt_num(r['delta_abs'])} |")
        L.append(f"| Δ %      | {(r['delta_pct']*100):.1f}% " if r['delta_pct'] is not None else "| Δ %      | n/a |")
        # formula trace (best-effort: show first matching formula touching this measure)
        if formula_map is not None and not formula_map.empty:
            hits = formula_map[formula_map["formula"].str.contains(r["measure"], case=False, na=False)]
            if not hits.empty:
                ex = hits.iloc[0]
                L.append(f"\n**Excel formula sample** ({ex['sheet']}!{ex['cell']}):  ")
                L.append(f"```\n{ex['formula']}\n```")
        L.append(f"\n{RESOLUTION_HEADING}")
        L.append("_(empty — to be filled by SME: fix-Excel / accept / fix-pipeline)_\n")
        L.append("---\n")
    return "\n".join(L)


# ---------- expert-resolution preservation ----------

_RESOLUTION_BLOCK_RE = re.compile(
    re.escape(DIMENSION_KEY_PREFIX) + r"([0-9a-f]{12})" + re.escape(DIMENSION_KEY_SUFFIX)
    + r".*?" + re.escape(RESOLUTION_HEADING) + r"\n(.*?)(?=\n---|\Z)",
    re.S,
)


def merge_expert_resolutions(old_md: str, new_md: str) -> str:
    """For every dim-key found in old_md with a non-empty resolution body,
    splice that body into new_md under the same dim-key."""
    if not old_md:
        return new_md
    old_resolutions: dict[str, str] = {}
    for m in _RESOLUTION_BLOCK_RE.finditer(old_md):
        key, body = m.group(1), m.group(2).strip()
        if body and not body.startswith("_(empty"):
            old_resolutions[key] = body
    if not old_resolutions:
        return new_md

    def _splice(m: re.Match) -> str:
        key = m.group(1)
        if key in old_resolutions:
            return (f"{DIMENSION_KEY_PREFIX}{key}{DIMENSION_KEY_SUFFIX}"
                    + m.group(0).split(DIMENSION_KEY_SUFFIX, 1)[1].split(RESOLUTION_HEADING)[0]
                    + RESOLUTION_HEADING + "\n" + old_resolutions[key] + "\n")
        return m.group(0)

    return _RESOLUTION_BLOCK_RE.sub(_splice, new_md)


# ---------- top-level entry ----------

def _to_pandas(df) -> pd.DataFrame:
    if isinstance(df, pd.DataFrame):
        return df
    # Spark DataFrame
    return df.toPandas()


def build_review(country_code: str,
                 excel_path,
                 pipeline_df,
                 dimension_cols: list[str],
                 value_cols: list[str],
                 excel_sheet: str = "Executed",
                 threshold: float = 0.05,
                 formula_map: pd.DataFrame | None = None,
                 out_path=None,
                 pipeline_source_ref: str = "") -> Path:
    excel_df = load_excel_totals(excel_path, excel_sheet, dimension_cols, value_cols)
    pdf = _to_pandas(pipeline_df)
    pdf = pdf.groupby(dimension_cols, dropna=False, as_index=False)[value_cols].sum(min_count=1)

    diff_df = _diff(excel_df, pdf, dimension_cols, value_cols, threshold)

    if formula_map is None:
        try:
            formula_map = load_formula_map(excel_path, target_sheets=(excel_sheet,))
        except Exception:
            formula_map = pd.DataFrame()

    if out_path is None:
        REVIEWS_DIR.mkdir(parents=True, exist_ok=True)
        out_path = REVIEWS_DIR / f"{country_code.lower()}_discrepancy_review.md"
    out_path = Path(out_path)

    new_md = _render(country_code, excel_path, dimension_cols, value_cols,
                     threshold, diff_df, formula_map, pipeline_source_ref)

    if out_path.exists():
        old = out_path.read_text()
        new_md = merge_expert_resolutions(old, new_md)

    out_path.write_text(new_md)
    return out_path
